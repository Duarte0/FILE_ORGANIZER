import os
import re
import shutil
import threading
import time
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


@dataclass
class RegrasCarregadas:
    empresas: dict
    rotas: dict
    palavras_chave: list
    total_empresas: int
    total_rotas: int
    total_palavras_chave: int


class ErroProcessamento(Exception):
    def __init__(self, etapa: str, motivo: str):
        super().__init__(motivo)
        self.etapa = etapa
        self.motivo = motivo


class DistribuidorArquivos:
    REGEX_NOME_ARQUIVO = re.compile(
        r"^\s*(?P<tipo>.+?)\s*-\s*(?P<empresa>.+?)\s*-\s*(?P<competencia>\d{2}\.\d{4})\.(?P<extensao>[^.]+)\s*$"
    )

    def __init__(self, base_dir: str, pasta_entrada: str, pasta_excel: str, pasta_relatorios: str, logger):
        self.base_dir = os.path.abspath(base_dir)
        self.logger = logger
        self.pasta_entrada = self._resolver_caminho(pasta_entrada)
        self.pasta_excel = self._resolver_caminho(pasta_excel or ".")
        self.pasta_relatorios = self._resolver_caminho(pasta_relatorios or "relatorios")
        self.arquivo_regras = os.path.join(self.pasta_excel, "regras.xlsx")
        self.pasta_nao_identificados = os.path.join(self.pasta_entrada, "_nao_identificados")
        self._lock_registros = threading.Lock()
        self._encerrado = False
        self._data_registros = datetime.now().date()
        self._registros = []

        os.makedirs(self.pasta_entrada, exist_ok=True)
        os.makedirs(self.pasta_relatorios, exist_ok=True)
        os.makedirs(self.pasta_nao_identificados, exist_ok=True)

    def _resolver_caminho(self, caminho: str) -> str:
        if not caminho:
            return self.base_dir

        caminho_expandido = os.path.expandvars(caminho.strip())

        if os.path.isabs(caminho_expandido):
            return os.path.normpath(caminho_expandido)

        return os.path.normpath(os.path.join(self.base_dir, caminho_expandido))

    def _normalizar_texto(self, valor) -> str:
        if valor is None:
            return ""

        texto = str(valor).strip().upper()
        return " ".join(texto.split())

    def _obter_indices_cabecalho(self, worksheet, obrigatorios: list[str]) -> dict:
        cabecalho = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

        if cabecalho is None:
            raise ErroProcessamento("Carga de regras", f"A aba {worksheet.title} esta vazia.")

        indices = {}

        for posicao, valor in enumerate(cabecalho):
            nome = self._normalizar_texto(valor)
            if nome:
                indices[nome] = posicao

        faltantes = [campo for campo in obrigatorios if campo not in indices]
        if faltantes:
            campos = ", ".join(faltantes)
            raise ErroProcessamento(
                "Carga de regras",
                f"A aba {worksheet.title} nao possui as colunas obrigatorias: {campos}.",
            )

        return indices

    def carregar_regras(self) -> RegrasCarregadas:
        # O Excel e relido a cada processamento para refletir alteracoes sem reiniciar o programa.
        if not os.path.exists(self.arquivo_regras):
            raise ErroProcessamento(
                "Carga de regras",
                f"Arquivo de regras nao encontrado em {self.arquivo_regras}.",
            )

        workbook = load_workbook(self.arquivo_regras, data_only=True)

        try:
            aba_empresas = workbook["EMPRESAS"]
            aba_rotas = workbook["ROTAS"]
            aba_palavras = workbook["PALAVRAS_CHAVE"]
        except KeyError as erro:
            workbook.close()
            raise ErroProcessamento("Carga de regras", f"Aba obrigatoria ausente no Excel: {erro}.") from erro

        indices_empresas = self._obter_indices_cabecalho(aba_empresas, ["EMPRESA", "CAMINHO_BASE"])
        indices_rotas = self._obter_indices_cabecalho(aba_rotas, ["TIPO_DOCUMENTO", "DEPARTAMENTO", "SUBPASTA"])
        indices_palavras = self._obter_indices_cabecalho(aba_palavras, ["PALAVRA_CHAVE", "TIPO_DOCUMENTO"])

        empresas = {}
        rotas = {}
        palavras_chave = []
        total_empresas = 0
        total_rotas = 0
        total_palavras = 0

        for linha in aba_empresas.iter_rows(min_row=2, values_only=True):
            empresa = linha[indices_empresas["EMPRESA"]]
            caminho_base = linha[indices_empresas["CAMINHO_BASE"]]

            empresa_normalizada = self._normalizar_texto(empresa)
            if not empresa_normalizada or not caminho_base:
                continue

            empresas[empresa_normalizada] = {
                "empresa": str(empresa).strip(),
                "caminho_base": os.path.normpath(str(caminho_base).strip()),
            }
            total_empresas += 1

        for linha in aba_rotas.iter_rows(min_row=2, values_only=True):
            tipo_documento = linha[indices_rotas["TIPO_DOCUMENTO"]]
            departamento = linha[indices_rotas["DEPARTAMENTO"]]
            subpasta = linha[indices_rotas["SUBPASTA"]]

            tipo_normalizado = self._normalizar_texto(tipo_documento)
            if not tipo_normalizado or not subpasta:
                continue

            if tipo_normalizado not in rotas:
                rotas[tipo_normalizado] = {
                    "tipo_documento": str(tipo_documento).strip(),
                    "departamento": "" if departamento is None else str(departamento).strip(),
                    "subpasta": str(subpasta).strip(),
                }

            total_rotas += 1

        for linha in aba_palavras.iter_rows(min_row=2, values_only=True):
            palavra = linha[indices_palavras["PALAVRA_CHAVE"]]
            tipo_documento = linha[indices_palavras["TIPO_DOCUMENTO"]]

            palavra_normalizada = self._normalizar_texto(palavra)
            tipo_normalizado = self._normalizar_texto(tipo_documento)
            if not palavra_normalizada or not tipo_normalizado:
                continue

            palavras_chave.append(
                {
                    "palavra_chave": palavra_normalizada,
                    "tipo_documento": str(tipo_documento).strip(),
                    "tipo_documento_normalizado": tipo_normalizado,
                }
            )
            total_palavras += 1

        palavras_chave.sort(key=lambda item: len(item["palavra_chave"]), reverse=True)
        workbook.close()

        return RegrasCarregadas(
            empresas=empresas,
            rotas=rotas,
            palavras_chave=palavras_chave,
            total_empresas=total_empresas,
            total_rotas=total_rotas,
            total_palavras_chave=total_palavras,
        )

    def obter_resumo_regras(self) -> RegrasCarregadas:
        return self.carregar_regras()

    def aguardar_arquivo_estavel(self, caminho_arquivo: str, tentativas: int = 120, intervalo: float = 0.5) -> None:
        # Aguarda o tamanho do arquivo se manter estavel para evitar processar copia incompleta.
        tamanho_anterior = None
        repeticoes_estaveis = 0

        for _ in range(tentativas):
            if not os.path.exists(caminho_arquivo):
                raise ErroProcessamento(
                    "Pre-processamento",
                    f"O arquivo {caminho_arquivo} nao esta mais disponivel para processamento.",
                )

            tamanho_atual = os.path.getsize(caminho_arquivo)

            if tamanho_atual == tamanho_anterior:
                repeticoes_estaveis += 1
            else:
                repeticoes_estaveis = 0

            if repeticoes_estaveis >= 1:
                return

            tamanho_anterior = tamanho_atual
            time.sleep(intervalo)

        raise ErroProcessamento(
            "Pre-processamento",
            f"O arquivo {caminho_arquivo} nao ficou estavel dentro do tempo esperado.",
        )

    def _extrair_dados_nome(self, nome_arquivo: str) -> dict:
        correspondencia = self.REGEX_NOME_ARQUIVO.match(nome_arquivo)
        if not correspondencia:
            raise ErroProcessamento(
                "Etapa 1 - Parse do nome",
                "O nome do arquivo nao segue o padrao {TIPO} - {EMPRESA} - {MM.AAAA}.{extensao}.",
            )

        dados = correspondencia.groupdict()
        mes, ano = dados["competencia"].split(".")

        return {
            "tipo_bruto": dados["tipo"].strip(),
            "empresa_bruta": dados["empresa"].strip(),
            "competencia": dados["competencia"],
            "mes": mes,
            "ano": ano,
            "extensao": dados["extensao"].strip(),
        }

    def _identificar_tipo_documento(self, tipo_bruto: str, palavras_chave: list[dict]) -> str:
        tipo_normalizado = self._normalizar_texto(tipo_bruto)

        for item in palavras_chave:
            if item["palavra_chave"] in tipo_normalizado:
                return item["tipo_documento"]

        raise ErroProcessamento(
            "Etapa 3 - Tipo nao identificado",
            f"Nenhuma palavra-chave cadastrada foi encontrada em '{tipo_bruto}'.",
        )

    def _montar_destino_sem_conflito(self, pasta_destino: str, nome_arquivo: str) -> str:
        destino = os.path.join(pasta_destino, nome_arquivo)

        if not os.path.exists(destino):
            return destino

        nome_base, extensao = os.path.splitext(nome_arquivo)
        contador = 1

        while True:
            candidato = os.path.join(pasta_destino, f"{nome_base}_{contador}{extensao}")
            if not os.path.exists(candidato):
                return candidato
            contador += 1

    def _rotacionar_relatorio_se_necessario(self, data_atual) -> None:
        with self._lock_registros:
            if self._data_registros != data_atual and self._registros:
                self._exportar_relatorio(self._data_registros, list(self._registros))
                self._registros.clear()
                self._data_registros = data_atual
            elif self._data_registros != data_atual:
                self._data_registros = data_atual

    def _registrar_resultado(self, registro: dict) -> None:
        data_registro = registro["timestamp"].date()
        self._rotacionar_relatorio_se_necessario(data_registro)

        with self._lock_registros:
            self._registros.append(registro)

        mensagem = (
            "arquivo=%s | empresa=%s | tipo=%s | etapa=%s | motivo=%s | destino=%s | status=%s"
        )
        argumentos = (
            registro["nome_arquivo"],
            registro["empresa_identificada"],
            registro["tipo_identificado"],
            registro["etapa"],
            registro["motivo"],
            registro["caminho_destino"],
            registro["status"],
        )

        if registro["status"] == "SUCESSO":
            self.logger.info(mensagem, *argumentos)
        else:
            self.logger.error(mensagem, *argumentos)

    def _exportar_relatorio(self, data_referencia, registros: list[dict]) -> str:
        # Gera um Excel diario com destaque visual para linhas com falha.
        os.makedirs(self.pasta_relatorios, exist_ok=True)
        caminho_relatorio = os.path.join(
            self.pasta_relatorios,
            f"relatorio_{data_referencia.isoformat()}.xlsx",
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Relatorio"

        cabecalhos = [
            "Timestamp",
            "Nome do arquivo",
            "Empresa identificada",
            "Tipo identificado",
            "Etapa",
            "Motivo",
            "Caminho de destino",
            "Status",
        ]
        worksheet.append(cabecalhos)

        for celula in worksheet[1]:
            celula.font = Font(bold=True)

        preenchimento_erro = PatternFill(fill_type="solid", fgColor="F4CCCC")

        for registro in registros:
            worksheet.append(
                [
                    registro["timestamp"].strftime("%Y-%m-%d %H:%M:%S"),
                    registro["nome_arquivo"],
                    registro["empresa_identificada"],
                    registro["tipo_identificado"],
                    registro["etapa"],
                    registro["motivo"],
                    registro["caminho_destino"],
                    registro["status"],
                ]
            )

            if registro["status"] == "ERRO":
                linha_atual = worksheet.max_row
                for celula in worksheet[linha_atual]:
                    celula.fill = preenchimento_erro

        for coluna in worksheet.columns:
            maior_tamanho = 0
            letra_coluna = coluna[0].column_letter
            for celula in coluna:
                valor = "" if celula.value is None else str(celula.value)
                maior_tamanho = max(maior_tamanho, len(valor))
            worksheet.column_dimensions[letra_coluna].width = min(maior_tamanho + 2, 80)

        workbook.save(caminho_relatorio)
        workbook.close()
        return caminho_relatorio

    def exportar_relatorio_diario(self) -> str | None:
        with self._lock_registros:
            if not self._registros:
                return None

            registros = list(self._registros)
            data_referencia = self._data_registros

        caminho_relatorio = self._exportar_relatorio(data_referencia, registros)
        self.logger.info("Relatorio exportado em %s", caminho_relatorio)
        return caminho_relatorio

    def _mover_para_nao_identificados(self, caminho_arquivo: str) -> str:
        os.makedirs(self.pasta_nao_identificados, exist_ok=True)
        nome_arquivo = os.path.basename(caminho_arquivo)
        destino = self._montar_destino_sem_conflito(self.pasta_nao_identificados, nome_arquivo)
        shutil.move(caminho_arquivo, destino)
        return destino

    def processar_arquivo(self, caminho_arquivo: str) -> dict:
        nome_arquivo = os.path.basename(caminho_arquivo)
        registro = {
            "timestamp": datetime.now(),
            "nome_arquivo": nome_arquivo,
            "empresa_identificada": "",
            "tipo_identificado": "",
            "etapa": "",
            "motivo": "",
            "caminho_destino": "",
            "status": "ERRO",
        }

        try:
            # Pre-processamento: aguarda a escrita terminar antes de abrir as regras e mover o arquivo.
            self.aguardar_arquivo_estavel(caminho_arquivo)
            regras = self.carregar_regras()

            # Etapa 1: extrai tipo bruto, empresa bruta e competencia a partir do nome do arquivo.
            dados_nome = self._extrair_dados_nome(nome_arquivo)
            empresa_normalizada = self._normalizar_texto(dados_nome["empresa_bruta"])

            # Etapa 2: identifica a empresa na aba EMPRESAS e recupera o caminho base.
            empresa = regras.empresas.get(empresa_normalizada)
            if not empresa:
                raise ErroProcessamento(
                    "Etapa 2 - Empresa nao cadastrada",
                    f"A empresa '{dados_nome['empresa_bruta']}' nao foi encontrada na aba EMPRESAS.",
                )

            registro["empresa_identificada"] = empresa["empresa"]

            # Etapa 3: encontra o tipo padronizado a partir das palavras-chave cadastradas.
            tipo_documento = self._identificar_tipo_documento(dados_nome["tipo_bruto"], regras.palavras_chave)
            registro["tipo_identificado"] = tipo_documento

            # Etapa 4: localiza a rota correspondente ao tipo de documento padronizado.
            rota = regras.rotas.get(self._normalizar_texto(tipo_documento))
            if not rota:
                raise ErroProcessamento(
                    "Etapa 4 - Rota nao cadastrada",
                    f"O tipo de documento '{tipo_documento}' nao possui rota cadastrada na aba ROTAS.",
                )

            # Etapa 5: monta o caminho final substituindo o placeholder {ANO} quando existir.
            subpasta = rota["subpasta"].replace("{ANO}", dados_nome["ano"])
            pasta_destino = os.path.normpath(os.path.join(empresa["caminho_base"], subpasta))

            try:
                # Etapa 6: cria as pastas intermediarias e move o arquivo para o destino final.
                os.makedirs(pasta_destino, exist_ok=True)

                destino_final = self._montar_destino_sem_conflito(pasta_destino, nome_arquivo)
                registro["caminho_destino"] = destino_final

                shutil.move(caminho_arquivo, destino_final)
            except Exception as erro_movimento:
                raise ErroProcessamento(
                    "Etapa 6 - Criacao de pasta ou movimentacao",
                    f"Falha ao criar a pasta de destino ou mover o arquivo: {erro_movimento}",
                ) from erro_movimento

            registro["etapa"] = "Concluido"
            registro["motivo"] = "Arquivo distribuido com sucesso."
            registro["status"] = "SUCESSO"
        except ErroProcessamento as erro:
            registro["etapa"] = erro.etapa
            registro["motivo"] = erro.motivo

            if os.path.exists(caminho_arquivo):
                try:
                    registro["caminho_destino"] = self._mover_para_nao_identificados(caminho_arquivo)
                except Exception as erro_movimento:
                    registro["motivo"] = f"{registro['motivo']} Falha ao mover para _nao_identificados: {erro_movimento}."
            elif not registro["caminho_destino"]:
                registro["caminho_destino"] = os.path.join(self.pasta_nao_identificados, nome_arquivo)
        except Exception as erro:
            registro["etapa"] = "Falha inesperada"
            registro["motivo"] = str(erro)

            if os.path.exists(caminho_arquivo):
                try:
                    registro["caminho_destino"] = self._mover_para_nao_identificados(caminho_arquivo)
                except Exception as erro_movimento:
                    registro["motivo"] = f"{registro['motivo']} Falha ao mover para _nao_identificados: {erro_movimento}."
            elif not registro["caminho_destino"]:
                registro["caminho_destino"] = os.path.join(self.pasta_nao_identificados, nome_arquivo)
        finally:
            registro["timestamp"] = datetime.now()
            self._registrar_resultado(registro)

        return registro

    def encerrar(self) -> None:
        if self._encerrado:
            return

        self._encerrado = True
        caminho_relatorio = self.exportar_relatorio_diario()
        if caminho_relatorio:
            self.logger.info("Encerramento concluido com relatorio salvo em %s", caminho_relatorio)
        else:
            self.logger.info("Encerramento concluido sem registros pendentes para exportacao.")